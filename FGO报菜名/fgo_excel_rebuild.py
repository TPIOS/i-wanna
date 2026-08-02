"""用已提取的图片重建带“置于单元格中”图片的 FGO 报名表。

依赖：
    python -m pip install -U XlsxWriter openpyxl Pillow

输入：
    看看龙大Master实力报名表_v3.xlsx   最新文字、表头和样式模板
    提取结果/                         职介/从者/御主.png

输出：
    看看龙大Master实力报名表_vFinal.xlsx

XlsxWriter 的 embed_image() 会生成 Excel 365 的“置于单元格中”图片。
统计公式同时写入按实际图片文件计算出的缓存值，因此打开文件时无需等待
Excel 重算，也能立即看到正确的 count。
"""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
import os
from pathlib import Path
import re
import zipfile

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
import xlsxwriter
from xlsxwriter.utility import xl_cell_to_rowcol, xl_col_to_name

from fgo_excel_extract import (
    WorksheetGrid,
    read_floating_images,
    read_shared_strings,
    read_worksheet,
    sheet_path,
)


# ===== 可以按需要修改的设置 =====
TEMPLATE_FILE = Path("看看龙大Master实力报名表_v3.xlsx")
IMAGE_ROOT = Path("提取结果")
OUTPUT_FILE = Path("看看龙大Master实力报名表_vFinal.xlsx")
SHEET_NAME: str | None = None  # None = 模板的第一个工作表

HEADER_ROW = 1
SERVANT_ROW = 2
FIRST_MASTER_ROW = 3
SUMMARY_ROW = 25
MASTER_NAME_COLUMN = 1  # A 列
PROFILE_IMAGE_COLUMN = 2  # B 列
FIRST_SERVANT_COLUMN = 5  # E 列
TOTAL_COLUMN = 195  # GM 列

# 同名输出存在时是否覆盖。模板文件本身永远不会被修改。
OVERWRITE_OUTPUT = True

# 将模板 B 列的御主资料截图也转换为单元格图片。
COPY_PROFILE_IMAGES = True

# 统计列，以及它统计的从者列范围。
COUNT_GROUPS = (
    ("Y", "E", "X"),
    ("AP", "Z", "AO"),
    ("BG", "AQ", "BF"),
    ("BZ", "BH", "BY"),
    ("CR", "CA", "CQ"),
    ("DF", "CS", "DE"),
    ("DX", "DG", "DW"),
    ("EL", "DY", "EK"),
    ("EW", "EM", "EV"),
    ("FD", "EX", "FC"),
    ("FQ", "FE", "FP"),
    ("GC", "FR", "GB"),
    ("GI", "GD", "GH"),
    ("GL", "GJ", "GK"),
)

SUPPORTED_IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".gif", ".bmp")


def safe_component(name: str) -> str:
    """使用与提取脚本相同的 Windows 文件名清理规则。"""
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name)).strip().rstrip(". ")
    if not cleaned:
        raise ValueError("文件或文件夹名称为空")
    if cleaned.upper() in {
        "CON", "PRN", "AUX", "NUL",
        *(f"COM{i}" for i in range(1, 10)),
        *(f"LPT{i}" for i in range(1, 10)),
    }:
        cleaned = f"_{cleaned}"
    return cleaned


def color_value(color) -> str | None:
    """将 openpyxl Color 转成 XlsxWriter 可用的 #RRGGBB。"""
    if color is None:
        return None

    color_type = getattr(color, "type", None)
    raw: str | None = None
    if color_type == "rgb":
        raw = getattr(color, "rgb", None)
    elif color_type == "indexed":
        index = getattr(color, "indexed", None)
        if isinstance(index, int) and 0 <= index < len(COLOR_INDEX):
            raw = COLOR_INDEX[index]

    if not raw or not isinstance(raw, str):
        return None
    raw = raw[-6:]
    return f"#{raw.upper()}" if re.fullmatch(r"[0-9A-Fa-f]{6}", raw) else None


BORDER_STYLES = {
    "thin": 1,
    "medium": 2,
    "dashed": 3,
    "dotted": 4,
    "thick": 5,
    "double": 6,
    "hair": 7,
    "mediumDashed": 8,
    "dashDot": 9,
    "mediumDashDot": 10,
    "dashDotDot": 11,
    "mediumDashDotDot": 12,
    "slantDashDot": 13,
}


def format_properties(cell) -> dict:
    """将模板单元格的常用样式转换成 XlsxWriter Format 属性。"""
    properties: dict = {}
    font = getattr(cell, "font", None)
    if font:
        if font.name:
            properties["font_name"] = font.name
        if font.sz:
            properties["font_size"] = font.sz
        if font.bold:
            properties["bold"] = True
        if font.italic:
            properties["italic"] = True
        if font.strike:
            properties["font_strikeout"] = True
        if font.underline:
            properties["underline"] = 2 if font.underline == "double" else 1
        font_color = color_value(font.color)
        if font_color:
            properties["font_color"] = font_color

    fill = getattr(cell, "fill", None)
    if fill and fill.patternType == "solid":
        fill_color = color_value(fill.fgColor)
        if fill_color:
            properties["bg_color"] = fill_color
            properties["pattern"] = 1

    border = getattr(cell, "border", None)
    if border:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name)
            style = BORDER_STYLES.get(side.style)
            if style:
                properties[side_name] = style
            side_color = color_value(side.color)
            if side_color:
                properties[f"{side_name}_color"] = side_color

    alignment = getattr(cell, "alignment", None)
    if alignment:
        if alignment.horizontal:
            properties["align"] = alignment.horizontal
        if alignment.vertical:
            properties["valign"] = {
                "center": "vcenter",
                "justify": "vjustify",
                "distributed": "vdistributed",
            }.get(alignment.vertical, alignment.vertical)
        if alignment.wrap_text:
            properties["text_wrap"] = True
        if alignment.shrink_to_fit:
            properties["shrink"] = True
        if alignment.text_rotation:
            properties["rotation"] = alignment.text_rotation
        if alignment.indent:
            properties["indent"] = alignment.indent

    number_format = getattr(cell, "number_format", None)
    if number_format and number_format != "General":
        properties["num_format"] = number_format

    protection = getattr(cell, "protection", None)
    if protection:
        if protection.locked is False:
            properties["locked"] = False
        if protection.hidden:
            properties["hidden"] = True

    return properties


class FormatCache:
    def __init__(self, workbook: xlsxwriter.Workbook) -> None:
        self.workbook = workbook
        self.cache: dict[int, xlsxwriter.format.Format] = {}

    def get(self, cell):
        style_id = int(getattr(cell, "style_id", 0) or 0)
        if style_id not in self.cache:
            self.cache[style_id] = self.workbook.add_format(format_properties(cell))
        return self.cache[style_id]


class SourceTable:
    """一次性缓存小型有效区域，避免只读模式反复扫描 400MB 模板。"""

    def __init__(self, worksheet) -> None:
        self.cells: dict[tuple[int, int], object] = {}
        rows = worksheet.iter_rows(
            min_row=1,
            max_row=SUMMARY_ROW,
            min_col=1,
            max_col=TOTAL_COLUMN,
        )
        for row_number, row_cells in enumerate(rows, start=1):
            for column_number, cell in enumerate(row_cells, start=1):
                self.cells[(row_number, column_number)] = cell

    def cell(self, row: int, column: int):
        return self.cells[(row, column)]


def copy_dimensions(worksheet_xml, target_worksheet) -> None:
    """复制默认行高、各列宽、特殊行高、隐藏状态和冻结窗格。"""
    sheet_format = next(
        (element for element in worksheet_xml if element.tag.rsplit("}", 1)[-1] == "sheetFormatPr"),
        None,
    )
    if sheet_format is not None and sheet_format.get("defaultRowHeight"):
        target_worksheet.set_default_row(float(sheet_format.get("defaultRowHeight")))

    for element in worksheet_xml.iter():
        name = element.tag.rsplit("}", 1)[-1]
        if name == "col":
            first = int(element.get("min", "1")) - 1
            last = int(element.get("max", str(first + 1))) - 1
            width = float(element.get("width", "8.43"))
            options = {}
            if element.get("hidden") in {"1", "true", "True"}:
                options["hidden"] = True
            target_worksheet.set_column(first, last, width, None, options)
        elif name == "row" and element.get("r"):
            row = int(element.get("r")) - 1
            height = float(element.get("ht")) if element.get("ht") else None
            options = {}
            if element.get("hidden") in {"1", "true", "True"}:
                options["hidden"] = True
            if height is not None or options:
                target_worksheet.set_row(row, height, None, options)

    pane = next(
        (element for element in worksheet_xml.iter() if element.tag.rsplit("}", 1)[-1] == "pane"),
        None,
    )
    if pane is not None:
        x_split = int(float(pane.get("xSplit", "0")))
        y_split = int(float(pane.get("ySplit", "0")))
        if x_split or y_split:
            target_worksheet.freeze_panes(y_split, x_split)


def merged_ranges(worksheet_xml) -> list[str]:
    return [
        element.get("ref")
        for element in worksheet_xml.iter()
        if element.tag.rsplit("}", 1)[-1] == "mergeCell" and element.get("ref")
    ]


def range_cells(reference: str) -> set[tuple[int, int]]:
    first, last = reference.split(":")
    first_row, first_col = xl_cell_to_rowcol(first)
    last_row, last_col = xl_cell_to_rowcol(last)
    return {
        (row, col)
        for row in range(first_row, last_row + 1)
        for col in range(first_col, last_col + 1)
    }


def write_source_value(worksheet, row: int, col: int, cell, cell_format) -> None:
    """按原始类型写入模板值，避免把普通文本误判为公式。"""
    value = cell.value
    if value is None:
        if int(getattr(cell, "style_id", 0) or 0):
            worksheet.write_blank(row, col, None, cell_format)
        return

    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        formula = str(value)
        if formula.startswith("=="):
            formula = formula[1:]
        elif not formula.startswith("="):
            formula = f"={formula}"
        worksheet.write_formula(row, col, formula, cell_format)
    elif isinstance(value, bool):
        worksheet.write_boolean(row, col, value, cell_format)
    elif isinstance(value, (int, float)):
        worksheet.write_number(row, col, value, cell_format)
    elif isinstance(value, (datetime, date, time)):
        worksheet.write_datetime(row, col, value, cell_format)
    else:
        worksheet.write_string(row, col, str(value), cell_format)


def find_master_rows(source_worksheet) -> list[int]:
    rows: list[int] = []
    for row in range(FIRST_MASTER_ROW, SUMMARY_ROW):
        value = source_worksheet.cell(row=row, column=MASTER_NAME_COLUMN).value
        if value not in (None, ""):
            rows.append(row)
    if not rows:
        raise ValueError("模板 A3:A24 中没有找到御主名称。")
    return rows


def find_servants(source_worksheet) -> list[tuple[int, str, str]]:
    count_columns = {xl_cell_to_rowcol(f"{count}1")[1] + 1 for count, _, _ in COUNT_GROUPS}
    current_class: str | None = None
    servants: list[tuple[int, str, str]] = []

    for column in range(FIRST_SERVANT_COLUMN, TOTAL_COLUMN):
        if column in count_columns:
            current_class = None
            continue

        header = source_worksheet.cell(row=HEADER_ROW, column=column).value
        if header not in (None, ""):
            current_class = str(header).strip()

        servant = source_worksheet.cell(row=SERVANT_ROW, column=column).value
        if current_class and servant not in (None, ""):
            servants.append((column, current_class, str(servant).strip()))

    if not servants:
        raise ValueError("模板 E2:GL2 中没有找到从者。")
    return servants


def locate_matrix_images(
    master_rows: list[int], servants: list[tuple[int, str, str]], source_worksheet
) -> tuple[dict[tuple[int, int], Path], set[Path]]:
    """按照“职介/从者/清理后的御主名.png”建立单元格图片表。"""
    assignments: dict[tuple[int, int], Path] = {}
    used: set[Path] = set()

    for column, class_name, servant_name in servants:
        folder = IMAGE_ROOT / safe_component(class_name) / safe_component(servant_name)
        for row in master_rows:
            master_name = source_worksheet.cell(row=row, column=MASTER_NAME_COLUMN).value
            stem = safe_component(str(master_name))
            image_path = next(
                (
                    folder / f"{stem}{extension}"
                    for extension in SUPPORTED_IMAGE_EXTENSIONS
                    if (folder / f"{stem}{extension}").is_file()
                ),
                None,
            )
            if image_path is not None:
                assignments[(row, column)] = image_path
                used.add(image_path.resolve())

    return assignments, used


def load_profile_images(master_rows: list[int]) -> dict[int, tuple[str, bytes]]:
    """从 v3 模板的浮动图片中选取 B 列每行最大的资料截图。"""
    if not COPY_PROFILE_IMAGES:
        return {}

    with zipfile.ZipFile(TEMPLATE_FILE) as archive:
        worksheet_path = sheet_path(archive, SHEET_NAME)
        shared_strings = read_shared_strings(archive)
        worksheet_xml, _, _ = read_worksheet(archive, worksheet_path, shared_strings)
        grid = WorksheetGrid.from_worksheet(worksheet_xml)
        floating_images = read_floating_images(archive, worksheet_path, grid)

        candidates: dict[int, list[tuple[str, bytes]]] = {}
        valid_rows = set(master_rows)
        for image in floating_images:
            if image.column == PROFILE_IMAGE_COLUMN and image.row in valid_rows:
                candidates.setdefault(image.row, []).append(
                    (image.internal_path, archive.read(image.internal_path))
                )

    # 一个单元格只能放一张图片；若同一行有多张，保留文件体积最大的一张。
    return {
        row: max(row_images, key=lambda item: len(item[1]))
        for row, row_images in candidates.items()
    }


def formula_format(source_worksheet, formats: FormatCache, row: int, column: int):
    return formats.get(source_worksheet.cell(row=row, column=column))


def write_statistics(
    worksheet,
    source_worksheet,
    formats: FormatCache,
    master_rows: list[int],
    servants: list[tuple[int, str, str]],
    matrix_images: dict[tuple[int, int], Path],
    profile_images: dict[int, tuple[str, bytes]],
) -> None:
    """重写 Y/AP/...、GM 以及第 25 行，并写入正确缓存值。"""
    count_columns: list[int] = []

    for row in master_rows:
        group_values: list[int] = []
        for count_name, start_name, end_name in COUNT_GROUPS:
            count_col = xl_cell_to_rowcol(f"{count_name}1")[1] + 1
            start_col = xl_cell_to_rowcol(f"{start_name}1")[1] + 1
            end_col = xl_cell_to_rowcol(f"{end_name}1")[1] + 1
            value = sum((row, column) in matrix_images for column in range(start_col, end_col + 1))
            formula = f"=COUNTA({start_name}{row}:{end_name}{row})"
            worksheet.write_formula(
                row - 1,
                count_col - 1,
                formula,
                formula_format(source_worksheet, formats, row, count_col),
                value,
            )
            group_values.append(value)
            if count_col not in count_columns:
                count_columns.append(count_col)

        total_formula = "=" + "+".join(
            f"{xl_col_to_name(column - 1)}{row}" for column in count_columns
        )
        worksheet.write_formula(
            row - 1,
            TOTAL_COLUMN - 1,
            total_formula,
            formula_format(source_worksheet, formats, row, TOTAL_COLUMN),
            sum(group_values),
        )

    # 第 25 行：A/B/C 基础统计。
    text_columns = (
        (MASTER_NAME_COLUMN, len(master_rows)),
        (PROFILE_IMAGE_COLUMN, len(profile_images)),
        (
            3,
            sum(
                source_worksheet.cell(row=row, column=3).value not in (None, "")
                for row in master_rows
            ),
        ),
    )
    first_master = min(master_rows)
    last_master = max(master_rows)
    for column, value in text_columns:
        name = xl_col_to_name(column - 1)
        worksheet.write_formula(
            SUMMARY_ROW - 1,
            column - 1,
            f"=COUNTA({name}{first_master}:{name}{last_master})",
            formula_format(source_worksheet, formats, SUMMARY_ROW, column),
            value,
        )

    servant_columns = {column for column, _, _ in servants}
    for column in servant_columns:
        name = xl_col_to_name(column - 1)
        value = sum((row, column) in matrix_images for row in master_rows)
        worksheet.write_formula(
            SUMMARY_ROW - 1,
            column - 1,
            f"=COUNTA({name}{first_master}:{name}{last_master})",
            formula_format(source_worksheet, formats, SUMMARY_ROW, column),
            value,
        )

    # 统计列和 GM 每个御主行都有公式，所以第 25 行结果等于御主数量。
    for column in (*count_columns, TOTAL_COLUMN):
        name = xl_col_to_name(column - 1)
        worksheet.write_formula(
            SUMMARY_ROW - 1,
            column - 1,
            f"=COUNTA({name}{first_master}:{name}{last_master})",
            formula_format(source_worksheet, formats, SUMMARY_ROW, column),
            len(master_rows),
        )


def copy_template_cells(
    worksheet,
    source_worksheet,
    formats: FormatCache,
    worksheet_xml,
) -> None:
    """复制模板文本/数值；矩阵和第 25 行由后续步骤重建。"""
    merged = merged_ranges(worksheet_xml)
    merged_members: set[tuple[int, int]] = set()

    for reference in merged:
        members = range_cells(reference)
        merged_members.update(members)
        top_row, top_col = min(members)
        cell = source_worksheet.cell(row=top_row + 1, column=top_col + 1)
        worksheet.merge_range(reference, cell.value, formats.get(cell))

    for row in range(1, SUMMARY_ROW + 1):
        for column in range(1, TOTAL_COLUMN + 1):
            zero_based = (row - 1, column - 1)
            if zero_based in merged_members:
                continue
            if FIRST_MASTER_ROW <= row <= SUMMARY_ROW and column >= FIRST_SERVANT_COLUMN:
                continue
            if row == SUMMARY_ROW:
                continue
            cell = source_worksheet.cell(row=row, column=column)
            write_source_value(worksheet, row - 1, column - 1, cell, formats.get(cell))


def all_extracted_images() -> set[Path]:
    return {
        path.resolve()
        for path in IMAGE_ROOT.rglob("*")
        if path.is_file() and path.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
    }


def validate_inputs() -> None:
    if not TEMPLATE_FILE.is_file():
        raise FileNotFoundError(f"找不到模板：{TEMPLATE_FILE.resolve()}")
    if not IMAGE_ROOT.is_dir():
        raise FileNotFoundError(f"找不到图片目录：{IMAGE_ROOT.resolve()}")
    if OUTPUT_FILE.resolve() == TEMPLATE_FILE.resolve():
        raise ValueError("OUTPUT_FILE 不能与 TEMPLATE_FILE 相同。")
    if OUTPUT_FILE.exists() and not OVERWRITE_OUTPUT:
        raise FileExistsError(f"输出文件已存在：{OUTPUT_FILE.resolve()}")


def main() -> None:
    validate_inputs()

    source_workbook = load_workbook(
        TEMPLATE_FILE,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        selected_sheet_name = SHEET_NAME or source_workbook.sheetnames[0]
        source_worksheet = SourceTable(source_workbook[selected_sheet_name])
        master_rows = find_master_rows(source_worksheet)
        servants = find_servants(source_worksheet)
        matrix_images, used_images = locate_matrix_images(master_rows, servants, source_worksheet)
        profile_images = load_profile_images(master_rows)

        with zipfile.ZipFile(TEMPLATE_FILE) as archive:
            worksheet_path = sheet_path(archive, selected_sheet_name)
            shared_strings = read_shared_strings(archive)
            worksheet_xml, _, _ = read_worksheet(archive, worksheet_path, shared_strings)

        temporary_file = OUTPUT_FILE.with_name(f"{OUTPUT_FILE.stem}.building.xlsx")
        if temporary_file.exists():
            temporary_file.unlink()

        # 将打包临时文件放在当前项目目录，避免某些 Conda/权限环境无法访问系统 Temp。
        workbook = xlsxwriter.Workbook(temporary_file, {"tmpdir": "."})
        try:
            worksheet = workbook.add_worksheet(selected_sheet_name)
            if not hasattr(worksheet, "embed_image"):
                raise RuntimeError("XlsxWriter 版本过旧，请运行：python -m pip install -U XlsxWriter")
            workbook.set_calc_mode("auto")
            formats = FormatCache(workbook)

            copy_dimensions(worksheet_xml, worksheet)
            copy_template_cells(worksheet, source_worksheet, formats, worksheet_xml)

            for (row, column), image_path in matrix_images.items():
                cell_format = formats.get(source_worksheet.cell(row=row, column=column))
                worksheet.embed_image(
                    row - 1,
                    column - 1,
                    str(image_path),
                    {"cell_format": cell_format, "description": image_path.stem},
                )

            for row, (internal_name, image_bytes) in profile_images.items():
                cell_format = formats.get(
                    source_worksheet.cell(row=row, column=PROFILE_IMAGE_COLUMN)
                )
                worksheet.embed_image(
                    row - 1,
                    PROFILE_IMAGE_COLUMN - 1,
                    internal_name,
                    {
                        "image_data": BytesIO(image_bytes),
                        "cell_format": cell_format,
                        "description": "御主资料截图",
                    },
                )

            write_statistics(
                worksheet,
                source_worksheet,
                formats,
                master_rows,
                servants,
                matrix_images,
                profile_images,
            )
        finally:
            workbook.close()

        if OUTPUT_FILE.exists():
            OUTPUT_FILE.unlink()
        os.replace(temporary_file, OUTPUT_FILE)

        extracted = all_extracted_images()
        unused = extracted - used_images
        print(f"御主：{len(master_rows)} 位")
        print(f"从者：{len(servants)} 名")
        print(f"写入从者图片：{len(matrix_images)} 张")
        print(f"写入 B 列资料图片：{len(profile_images)} 张")
        if unused:
            print(f"警告：提取目录中有 {len(unused)} 张图片未匹配到模板单元格。")
            for path in sorted(unused)[:10]:
                print(f"  - {path}")
        print(f"完成：{OUTPUT_FILE.resolve()}")
    finally:
        source_workbook.close()


if __name__ == "__main__":
    main()
