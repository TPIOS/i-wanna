"""从 FGO 御主报名表中提取各职介的持有率与御主统计。"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


COUNT_GROUPS = (
    ("Y", "E", "X"),
    ("AP", "Z", "AO"),
    ("BG", "AQ", "BF"),
    ("BZ", "BH", "BY"),
    ("CR", "CA", "CQ"),
    ("DF", "CS", "DE"),
    ("DX", "DG", "DW"),
    ("EL", "DY", "EK"),
    ("EW", "EM", "EV"),
    ("FD", "EX", "FC"),
    ("FQ", "FE", "FP"),
    ("GC", "FR", "GB"),
    ("GI", "GD", "GH"),
    ("GL", "GJ", "GK"),
)


def number(value: object, cell: str) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    raise ValueError(f"{cell} 没有可用的缓存统计值：{value!r}")


def analyze(workbook_path: Path) -> dict:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        master_rows = [
            row for row in range(3, 25) if worksheet.cell(row=row, column=1).value not in (None, "")
        ]
        masters = {
            row: str(worksheet.cell(row=row, column=1).value).strip() for row in master_rows
        }

        classes: list[dict] = []
        for count_name, start_name, end_name in COUNT_GROUPS:
            count_col = column_index_from_string(count_name)
            start_col = column_index_from_string(start_name)
            end_col = column_index_from_string(end_name)
            class_name = str(worksheet.cell(row=1, column=start_col).value).strip()

            servants: list[dict] = []
            for column in range(start_col, end_col + 1):
                servant = worksheet.cell(row=2, column=column).value
                if servant in (None, ""):
                    continue
                count = number(worksheet.cell(row=25, column=column).value, f"row25/col{column}")
                servants.append(
                    {
                        "name": str(servant).strip(),
                        "holding_count": count,
                        "holding_rate": count / len(master_rows),
                        "implementation_order": column,
                    }
                )
            if not servants:
                raise ValueError(f"{class_name} 没有找到从者")

            # 同持有数时，列号更小者即工作簿中更早实装的从者。
            highest = min(servants, key=lambda item: (-item["holding_count"], item["implementation_order"]))
            lowest = min(servants, key=lambda item: (item["holding_count"], item["implementation_order"]))

            master_counts = []
            for row in master_rows:
                count = number(worksheet.cell(row=row, column=count_col).value, f"{count_name}{row}")
                master_counts.append({"name": masters[row], "count": count, "row": row})
            best_count = max(item["count"] for item in master_counts)
            best_masters = [item["name"] for item in master_counts if item["count"] == best_count]

            classes.append(
                {
                    "class": class_name,
                    "servant_count": len(servants),
                    "master_count": len(master_rows),
                    "highest": highest,
                    "lowest": lowest,
                    "best_master_count": best_count,
                    "best_masters": best_masters,
                    "servants": servants,
                }
            )

        return {
            "source": str(workbook_path.resolve()),
            "sheet": worksheet.title,
            "master_count": len(master_rows),
            "classes": classes,
        }
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=Path("职介统计.json"))
    args = parser.parse_args()

    result = analyze(args.workbook)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    for item in result["classes"]:
        summary = (
            f"{item['class']}: 最高 {item['highest']['name']} "
            f"{item['highest']['holding_count']}/{item['master_count']}；"
            f"最低 {item['lowest']['name']} {item['lowest']['holding_count']}/{item['master_count']}；"
            f"最契合御主 {item['best_master_count']}/{item['servant_count']} "
            f"{'、'.join(item['best_masters'])}"
        )
        print(summary.encode("unicode_escape").decode("ascii"))
    print(f"已写入：{args.output.resolve()}")


if __name__ == "__main__":
    main()
