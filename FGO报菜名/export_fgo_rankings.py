"""从修正后的 FGO 报名表导出从者与御主持有量排名工作簿。"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string


COUNT_GROUPS = (
    ("Y", "E", "X"),
    ("AP", "Z", "AO"),
    ("BG", "AQ", "BF"),
    ("BZ", "BH", "BY"),
    ("CR", "CA", "CQ"),
    ("DF", "CS", "DE"),
    ("DX", "DG", "DW"),
    ("EL", "DY", "EK"),
    ("EW", "EM", "EV"),
    ("FD", "EX", "FC"),
    ("FQ", "FE", "FP"),
    ("GC", "FR", "GB"),
    ("GI", "GD", "GH"),
    ("GL", "GJ", "GK"),
)

FIRST_MASTER_ROW = 3
LAST_MASTER_ROW = 24
SUMMARY_ROW = 25
MASTER_NAME_COLUMN = 1
TOTAL_COLUMN = 195  # GM


@dataclass(frozen=True)
class ServantRecord:
    name: str
    count: int
    source_order: int
    class_name: str


@dataclass(frozen=True)
class MasterRecord:
    name: str
    total: int
    source_order: int
    class_counts: dict[str, int]


@dataclass(frozen=True)
class ClassRecord:
    name: str
    count_column: int
    servant_count: int
    servants: tuple[ServantRecord, ...]


def cached_integer(value: object, cell_reference: str) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)) and int(value) == value:
        return int(value)
    raise ValueError(f"{cell_reference} 没有正确的整数缓存值：{value!r}")


def read_statistics(source: Path) -> tuple[list[ServantRecord], list[MasterRecord], list[ClassRecord]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        master_rows = [
            row
            for row in range(FIRST_MASTER_ROW, LAST_MASTER_ROW + 1)
            if worksheet.cell(row=row, column=MASTER_NAME_COLUMN).value not in (None, "")
        ]
        if len(master_rows) != 20:
            raise ValueError(f"预期 20 位御主，实际读取到 {len(master_rows)} 位")

        classes: list[ClassRecord] = []
        servants: list[ServantRecord] = []
        for count_name, start_name, end_name in COUNT_GROUPS:
            count_column = column_index_from_string(count_name)
            start_column = column_index_from_string(start_name)
            end_column = column_index_from_string(end_name)
            class_name_value = worksheet.cell(row=1, column=start_column).value
            if class_name_value in (None, ""):
                raise ValueError(f"{start_name}1 缺少职介名称")
            class_name = str(class_name_value).strip()

            class_servants: list[ServantRecord] = []
            for column in range(start_column, end_column + 1):
                servant_value = worksheet.cell(row=2, column=column).value
                if servant_value in (None, ""):
                    continue
                count = cached_integer(
                    worksheet.cell(row=SUMMARY_ROW, column=column).value,
                    f"row{SUMMARY_ROW}/column{column}",
                )
                if not 0 <= count <= len(master_rows):
                    raise ValueError(f"从者持有量超出范围：{servant_value} = {count}")
                record = ServantRecord(
                    name=str(servant_value).strip(),
                    count=count,
                    source_order=column,
                    class_name=class_name,
                )
                class_servants.append(record)
                servants.append(record)

            if not class_servants:
                raise ValueError(f"{class_name} 没有从者")
            classes.append(
                ClassRecord(
                    name=class_name,
                    count_column=count_column,
                    servant_count=len(class_servants),
                    servants=tuple(class_servants),
                )
            )

        if len(servants) != 176:
            raise ValueError(f"预期 176 名从者，实际读取到 {len(servants)} 名")

        masters: list[MasterRecord] = []
        for row in master_rows:
            name = str(worksheet.cell(row=row, column=MASTER_NAME_COLUMN).value).strip()
            class_counts: dict[str, int] = {}
            for class_record in classes:
                count = cached_integer(
                    worksheet.cell(row=row, column=class_record.count_column).value,
                    f"row{row}/column{class_record.count_column}",
                )
                if not 0 <= count <= class_record.servant_count:
                    raise ValueError(f"御主职介持有量超出范围：{name}/{class_record.name} = {count}")
                class_counts[class_record.name] = count

            total = cached_integer(worksheet.cell(row=row, column=TOTAL_COLUMN).value, f"GM{row}")
            calculated_total = sum(class_counts.values())
            if total != calculated_total:
                raise ValueError(f"御主总数与职介之和不一致：{name}，GM={total}，职介合计={calculated_total}")
            masters.append(
                MasterRecord(
                    name=name,
                    total=total,
                    source_order=row,
                    class_counts=class_counts,
                )
            )

        servant_total = sum(record.count for record in servants)
        master_total = sum(record.total for record in masters)
        if servant_total != master_total:
            raise ValueError(f"双向总量不一致：从者列合计={servant_total}，御主行合计={master_total}")

        for class_record in classes:
            servant_class_total = sum(record.count for record in class_record.servants)
            master_class_total = sum(record.class_counts[class_record.name] for record in masters)
            if servant_class_total != master_class_total:
                raise ValueError(
                    f"{class_record.name} 双向总量不一致："
                    f"从者列={servant_class_total}，御主行={master_class_total}"
                )

        return servants, masters, classes
    finally:
        workbook.close()


HEADER_FILL = PatternFill("solid", fgColor="203864")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10, color="1F1F1F")
ALT_FILL = PatternFill("solid", fgColor="EAF2F8")
THIN_BLUE = Side(style="thin", color="9EBCD4")
BODY_BORDER = Border(bottom=THIN_BLUE)


def format_ranking_sheet(worksheet, denominator: int, row_count: int) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:B{row_count + 1}"
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 18
    worksheet.row_dimensions[1].height = 24
    worksheet.sheet_view.showGridLines = False

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, row_count + 2):
        worksheet.row_dimensions[row].height = 21
        for column in range(1, 3):
            cell = worksheet.cell(row=row, column=column)
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "center",
                vertical="center",
            )
            if row % 2 == 0:
                cell.fill = ALT_FILL
        worksheet.cell(row=row, column=2).number_format = f'0"/{denominator}"'

    worksheet.conditional_formatting.add(
        f"B2:B{row_count + 1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=denominator, color="5B9BD5"),
    )
    worksheet.print_area = f"A1:B{row_count + 1}"
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddHeader.center.text = worksheet.title
    worksheet.oddFooter.center.text = "第 &P / &N 页"


def write_ranking_sheet(worksheet, rows: list[tuple[str, int]], denominator: int, name_header: str) -> None:
    worksheet.append([name_header, "持有量"])
    for name, count in rows:
        worksheet.append([name, count])
    format_ranking_sheet(worksheet, denominator, len(rows))


def save_servant_workbook(output: Path, servants: list[ServantRecord], master_count: int) -> None:
    workbook = Workbook()
    try:
        workbook.properties.title = "FGO 从者持有率排名"
        workbook.properties.creator = "OpenAI Codex"
        worksheet = workbook.active
        worksheet.title = "从者持有率排名"
        ranked = sorted(servants, key=lambda record: (-record.count, record.source_order))
        write_ranking_sheet(
            worksheet,
            [(record.name, record.count) for record in ranked],
            master_count,
            "从者名",
        )
        workbook.save(output)
    finally:
        workbook.close()


def save_master_workbook(output: Path, masters: list[MasterRecord], classes: list[ClassRecord]) -> None:
    workbook = Workbook()
    try:
        workbook.properties.title = "FGO 御主持有从者数排名"
        workbook.properties.creator = "OpenAI Codex"

        combined = workbook.active
        combined.title = "综合数据"
        ranked_total = sorted(masters, key=lambda record: (-record.total, record.source_order))
        write_ranking_sheet(
            combined,
            [(record.name, record.total) for record in ranked_total],
            sum(class_record.servant_count for class_record in classes),
            "御主名",
        )

        for class_record in classes:
            worksheet = workbook.create_sheet(class_record.name)
            ranked = sorted(
                masters,
                key=lambda record: (-record.class_counts[class_record.name], record.source_order),
            )
            write_ranking_sheet(
                worksheet,
                [(record.name, record.class_counts[class_record.name]) for record in ranked],
                class_record.servant_count,
                "御主名",
            )

        workbook.save(output)
    finally:
        workbook.close()


def validate_output(
    servant_output: Path,
    master_output: Path,
    servant_count: int,
    master_count: int,
    classes: list[ClassRecord],
) -> None:
    servant_workbook = load_workbook(servant_output, read_only=True, data_only=True)
    try:
        if servant_workbook.sheetnames != ["从者持有率排名"]:
            raise ValueError(f"从者排名工作簿 Sheet 异常：{servant_workbook.sheetnames}")
        worksheet = servant_workbook[servant_workbook.sheetnames[0]]
        if worksheet.max_row != servant_count + 1 or worksheet.max_column != 2:
            raise ValueError("从者排名工作簿行列数异常")
        if worksheet["A1"].value != "从者名" or worksheet["B1"].value != "持有量":
            raise ValueError("从者排名表头异常")
        if worksheet["B2"].number_format != f'0"/{master_count}"':
            raise ValueError(f"从者排名显示分母异常：{worksheet['B2'].number_format}")
        values = [worksheet.cell(row=row, column=2).value for row in range(2, worksheet.max_row + 1)]
        if values != sorted(values, reverse=True):
            raise ValueError("从者排名没有按持有量降序排列")
    finally:
        servant_workbook.close()

    master_workbook = load_workbook(master_output, read_only=True, data_only=True)
    try:
        expected_sheets = ["综合数据", *(record.name for record in classes)]
        if master_workbook.sheetnames != expected_sheets:
            raise ValueError(f"御主排名工作簿 Sheet 异常：{master_workbook.sheetnames}")
        denominators = {"综合数据": servant_count}
        denominators.update({record.name: record.servant_count for record in classes})
        for sheet_name in expected_sheets:
            worksheet = master_workbook[sheet_name]
            if worksheet.max_row != master_count + 1 or worksheet.max_column != 2:
                raise ValueError(f"{sheet_name} 行列数异常")
            if worksheet["A1"].value != "御主名" or worksheet["B1"].value != "持有量":
                raise ValueError(f"{sheet_name} 表头异常")
            expected_format = f'0"/{denominators[sheet_name]}"'
            if worksheet["B2"].number_format != expected_format:
                raise ValueError(
                    f"{sheet_name} 显示分母异常：{worksheet['B2'].number_format} != {expected_format}"
                )
            values = [worksheet.cell(row=row, column=2).value for row in range(2, worksheet.max_row + 1)]
            if values != sorted(values, reverse=True):
                raise ValueError(f"{sheet_name} 没有按持有量降序排列")
    finally:
        master_workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--servant-output", type=Path, default=Path("从者持有率排名.xlsx"))
    parser.add_argument("--master-output", type=Path, default=Path("御主持有从者数排名.xlsx"))
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    source = args.source.resolve()
    servant_output = args.servant_output.resolve()
    master_output = args.master_output.resolve()
    for output in (servant_output, master_output):
        if output.exists() and not args.force:
            raise FileExistsError(f"输出文件已存在；如需覆盖请加 --force：{output}")
        if output == source:
            raise ValueError("输出文件不能覆盖源工作簿")

    servants, masters, classes = read_statistics(source)
    save_servant_workbook(servant_output, servants, len(masters))
    save_master_workbook(master_output, masters, classes)
    validate_output(
        servant_output,
        master_output,
        len(servants),
        len(masters),
        classes,
    )

    print(f"从者：{len(servants)} 名；御主：{len(masters)} 位；职介：{len(classes)} 个")
    print(f"双向持有记录总数：{sum(record.count for record in servants)}")
    top_servants = sorted(servants, key=lambda record: (-record.count, record.source_order))[:5]
    top_masters = sorted(masters, key=lambda record: (-record.total, record.source_order))[:5]
    servant_report = "；".join(f"{record.name} {record.count}/{len(masters)}" for record in top_servants)
    master_report = "；".join(f"{record.name} {record.total}/{len(servants)}" for record in top_masters)
    print(("从者前五：" + servant_report).encode("unicode_escape").decode("ascii"))
    print(("御主前五：" + master_report).encode("unicode_escape").decode("ascii"))
    print(f"从者排名：{servant_output}")
    print(f"御主排名：{master_output}")


if __name__ == "__main__":
    main()
