#!/usr/bin/env python3
"""从 Excel 读取 FGO 五星从者，并下载 fgo.wiki 的宝具相关语音。

默认输入：看看龙大Master实力报名表_vFinal2_nameCorrected.xlsx
默认输出：从者语音/<职介>/<从者名>/宝具卡1.mp3（或宝具1.mp3）

脚本只依赖 openpyxl；非 MP3 源文件使用当前 Conda 环境中的 ffmpeg
解码，并优先用 LAME 编码 MP3。网络请求、页面缓存、断点续跑和失败
报告均内置。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import random
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import BinaryIO, Callable, TypeVar
from urllib.error import HTTPError, URLError
from urllib.parse import quote, unquote, urljoin, urlsplit, urlunsplit
from urllib.request import Request, urlopen


DEFAULT_WORKBOOK = Path("看看龙大Master实力报名表_vFinal2_nameCorrected.xlsx")
DEFAULT_OUTPUT_DIR = Path("从者语音")
FGO_WIKI_BASE = "https://fgo.wiki/w/"

KNOWN_CLASSES = {
    "Shielder",
    "Saber",
    "Archer",
    "Lancer",
    "Rider",
    "Caster",
    "Assassin",
    "Berserker",
    "Ruler",
    "Avenger",
    "MoonCancer",
    "AlterEgo",
    "Foreigner",
    "Pretender",
    "Beast",
}

RETRYABLE_HTTP_CODES = {408, 425, 429, 500, 502, 503, 504}
INVALID_COMPONENT = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
VOICE_LABEL = re.compile(
    r"^(宝具卡|宝具)\s*(\d+)?(?:\s*[（(][^）)]*[）)])?$"
)
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0 Safari/537.36 FGO-Voice-Downloader/1.0"
)


@dataclass(frozen=True)
class Servant:
    column: str
    class_name: str
    name: str


@dataclass(frozen=True)
class VoiceClip:
    label: str
    base_name: str
    url: str
    filename: str


def configure_console() -> None:
    """在 Windows 下尽量让中文日志保持为 UTF-8。"""
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure is not None:
            reconfigure(encoding="utf-8", errors="replace")


def configure_logging(output_dir: Path, verbose: bool) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    level = logging.DEBUG if verbose else logging.INFO
    formatter = logging.Formatter(
        "%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S"
    )

    root = logging.getLogger()
    root.setLevel(level)
    root.handlers.clear()

    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(formatter)
    root.addHandler(console)

    file_handler = logging.FileHandler(
        output_dir / "download.log", mode="a", encoding="utf-8"
    )
    file_handler.setLevel(level)
    file_handler.setFormatter(formatter)
    root.addHandler(file_handler)


def normalise_text(value: object) -> str:
    if value is None:
        return ""
    # 从者正式名中的全角「：」「／」具有语义，不能用 NFKC 改成 ASCII。
    return re.sub(r"\s+", " ", str(value)).strip()


def normalise_label(value: object) -> str:
    """语音标签允许把全角数字、括号等转成统一形式。"""
    return unicodedata.normalize("NFKC", normalise_text(value))


def safe_component(value: str) -> str:
    value = INVALID_COMPONENT.sub("_", normalise_text(value)).rstrip(" .")
    if not value:
        return "未命名"
    reserved = {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        *(f"COM{i}" for i in range(1, 10)),
        *(f"LPT{i}" for i in range(1, 10)),
    }
    if value.upper() in reserved:
        value = f"_{value}"
    return value


def read_servants(
    workbook_path: Path, sheet_name: str | None, start_column: str
) -> list[Servant]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string, get_column_letter
    except ImportError as exc:  # pragma: no cover - friendly runtime error
        raise RuntimeError("缺少 openpyxl，请先在 fgo 环境中安装 openpyxl。") from exc

    if not workbook_path.is_file():
        raise FileNotFoundError(f"找不到 Excel：{workbook_path}")

    workbook = load_workbook(
        workbook_path, read_only=True, data_only=False, keep_links=False
    )
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = "、".join(workbook.sheetnames)
                raise ValueError(
                    f"找不到工作表 {sheet_name!r}；可用工作表：{available}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        start = column_index_from_string(start_column)
        rows = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=2,
                min_col=start,
                max_col=worksheet.max_column,
                values_only=True,
            )
        )
        if len(rows) < 2:
            raise ValueError("工作表不足两行，无法读取职介和从者名。")

        servants: list[Servant] = []
        current_class: str | None = None
        for offset, (header_value, name_value) in enumerate(zip(rows[0], rows[1])):
            header = normalise_text(header_value)
            if header in KNOWN_CLASSES:
                current_class = header

            name = normalise_text(name_value)
            if (
                current_class
                and name
                and not name.startswith("=")
                and name.lower() not in {"count", "total count"}
            ):
                servants.append(
                    Servant(
                        column=get_column_letter(start + offset),
                        class_name=current_class,
                        name=name,
                    )
                )

        duplicates = [
            name for name, count in Counter(s.name for s in servants).items() if count > 1
        ]
        if duplicates:
            raise ValueError(f"Excel 中存在重复从者名：{'、'.join(duplicates)}")
        if not servants:
            raise ValueError("没有从 Excel 第 1、2 行读取到从者。")
        return servants
    finally:
        workbook.close()


T = TypeVar("T")


class HttpClient:
    def __init__(self, timeout: float, retries: int, delay: float) -> None:
        self.timeout = timeout
        self.retries = retries
        self.delay = delay
        self._last_request = 0.0

    def _throttle(self) -> None:
        remaining = self.delay - (time.monotonic() - self._last_request)
        if remaining > 0:
            time.sleep(remaining)
        self._last_request = time.monotonic()

    @staticmethod
    def _retry_after(error: Exception) -> float | None:
        if isinstance(error, HTTPError):
            value = error.headers.get("Retry-After")
            if value and value.isdigit():
                return min(float(value), 120.0)
        return None

    @staticmethod
    def _can_retry(error: Exception) -> bool:
        if isinstance(error, HTTPError):
            return error.code in RETRYABLE_HTTP_CODES
        return isinstance(error, (URLError, TimeoutError, OSError))

    def _with_retries(self, url: str, operation: Callable[[], T]) -> T:
        for attempt in range(self.retries + 1):
            try:
                self._throttle()
                return operation()
            except Exception as exc:
                if attempt >= self.retries or not self._can_retry(exc):
                    raise
                wait = self._retry_after(exc)
                if wait is None:
                    wait = min(60.0, 2**attempt + random.uniform(0.25, 1.25))
                logging.warning(
                    "请求失败，%.1f 秒后重试 (%d/%d)：%s；%s",
                    wait,
                    attempt + 1,
                    self.retries,
                    url,
                    exc,
                )
                time.sleep(wait)
        raise AssertionError("unreachable")

    @staticmethod
    def _request(url: str, accept: str, referer: str | None = None) -> Request:
        headers = {
            "User-Agent": USER_AGENT,
            "Accept": accept,
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
            "Connection": "close",
        }
        if referer:
            headers["Referer"] = referer
        return Request(url, headers=headers)

    def get_bytes(self, url: str) -> bytes:
        def operation() -> bytes:
            request = self._request(url, "text/html,application/xhtml+xml,*/*;q=0.8")
            with urlopen(request, timeout=self.timeout) as response:
                return response.read()

        return self._with_retries(url, operation)

    def download(self, url: str, destination: Path, referer: str) -> str:
        destination.parent.mkdir(parents=True, exist_ok=True)
        partial = destination.with_name(destination.name + ".part")

        def operation() -> str:
            partial.unlink(missing_ok=True)
            request = self._request(url, "audio/*,application/octet-stream,*/*;q=0.5", referer)
            try:
                with urlopen(request, timeout=self.timeout) as response:
                    content_type = response.headers.get("Content-Type", "")
                    expected = response.headers.get("Content-Length")
                    written = self._copy_response(response, partial)
                if written == 0:
                    raise OSError("服务器返回了空文件")
                if expected and expected.isdigit() and written != int(expected):
                    raise OSError(
                        f"下载长度不完整：得到 {written}，预期 {expected}"
                    )
                with partial.open("rb") as source:
                    prefix = source.read(512).lstrip().lower()
                if prefix.startswith((b"<!doctype html", b"<html")):
                    raise OSError("音频地址返回了 HTML 页面")
                os.replace(partial, destination)
                return content_type
            except Exception:
                partial.unlink(missing_ok=True)
                raise

        return self._with_retries(url, operation)

    @staticmethod
    def _copy_response(response: BinaryIO, destination: Path) -> int:
        written = 0
        with destination.open("wb") as output:
            while True:
                chunk = response.read(1024 * 256)
                if not chunk:
                    break
                output.write(chunk)
                written += len(chunk)
        return written


class VoicePageParser(HTMLParser):
    """提取每个表格行的标题和 audio/source 下载地址。"""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[tuple[str, list[str]]] = []
        self._row: dict[str, list[str]] | None = None
        self._in_header = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        attributes = {key.lower(): value for key, value in attrs}
        if tag == "tr":
            self._finish_row()
            self._row = {
                "label": [],
                "source": [],
                "audio": [],
                "download": [],
            }
            self._in_header = 0
            return
        if self._row is None:
            return
        if tag == "th":
            self._in_header += 1
        elif tag == "source":
            url = attributes.get("src") or attributes.get("data-src")
            if url:
                self._row["source"].append(url)
        elif tag == "audio":
            url = attributes.get("src") or attributes.get("data-src")
            if url:
                self._row["audio"].append(url)
        elif tag == "a" and "download" in attributes:
            url = attributes.get("href")
            if url and not url.lower().startswith("javascript:"):
                self._row["download"].append(url)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "th" and self._in_header:
            self._in_header -= 1
        elif tag == "tr":
            self._finish_row()

    def handle_data(self, data: str) -> None:
        if self._row is not None and self._in_header:
            self._row["label"].append(data)

    def close(self) -> None:
        super().close()
        self._finish_row()

    def _finish_row(self) -> None:
        if self._row is None:
            return
        label = normalise_text("".join(self._row["label"]))
        urls = self._row["source"] or self._row["audio"] or self._row["download"]
        if label and urls:
            self.rows.append((label, list(urls)))
        self._row = None
        self._in_header = 0


def canonical_media_identity(url: str) -> str:
    parts = urlsplit(url)
    return urlunsplit((parts.scheme.lower(), parts.netloc.lower(), parts.path, "", ""))


def extract_voice_clips(page_html: str, page_url: str) -> list[VoiceClip]:
    parser = VoicePageParser()
    parser.feed(page_html)
    parser.close()

    raw_clips: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str]] = set()
    for raw_label, urls in parser.rows:
        label = normalise_label(raw_label)
        match = VOICE_LABEL.fullmatch(label)
        if not match:
            continue
        kind, number = match.groups()
        base_name = f"{kind}{number or '1'}"
        url = urljoin(page_url, urls[0])
        identity = (base_name, canonical_media_identity(url))
        if identity in seen:
            continue
        seen.add(identity)
        raw_clips.append((label, base_name, url))

    totals = Counter(base for _, base, _ in raw_clips)
    indexes: Counter[str] = Counter()
    clips: list[VoiceClip] = []
    for label, base_name, url in raw_clips:
        indexes[base_name] += 1
        suffix = "" if indexes[base_name] == 1 else f"_{indexes[base_name]}"
        filename = f"{base_name}{suffix}.mp3"
        clips.append(VoiceClip(label, base_name, url, filename))

    duplicate_labels = [name for name, count in totals.items() if count > 1]
    if duplicate_labels:
        logging.info(
            "同名语音存在多个不同媒体文件，已使用 _2、_3 后缀：%s",
            "、".join(duplicate_labels),
        )
    return clips


def voice_page_url(servant_name: str) -> str:
    return FGO_WIKI_BASE + quote(f"{servant_name}/语音", safe="/")


def page_cache_path(cache_dir: Path, servant_name: str, page_url: str) -> Path:
    digest = hashlib.sha256(page_url.encode("utf-8")).hexdigest()[:12]
    stem = safe_component(servant_name)[:60]
    return cache_dir / "pages" / f"{stem}-{digest}.html"


def load_page(
    client: HttpClient,
    cache_dir: Path,
    servant_name: str,
    page_url: str,
    refresh: bool,
) -> tuple[str, Path, bool]:
    cache_path = page_cache_path(cache_dir, servant_name, page_url)
    if cache_path.is_file() and cache_path.stat().st_size > 0 and not refresh:
        return cache_path.read_text(encoding="utf-8", errors="replace"), cache_path, True

    data = client.get_bytes(page_url)
    prefix = data[:1024].lower()
    if b"<html" not in prefix and b"<!doctype html" not in prefix:
        raise OSError(f"语音页不像 HTML：{page_url}")
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    partial = cache_path.with_name(cache_path.name + ".part")
    partial.write_bytes(data)
    os.replace(partial, cache_path)
    return data.decode("utf-8", errors="replace"), cache_path, False


def find_ffmpeg(explicit: str | None) -> Path:
    candidates: list[Path] = []
    if explicit:
        candidates.append(Path(explicit).expanduser())
    found = shutil.which("ffmpeg")
    if found:
        candidates.append(Path(found))
    if os.name == "nt":
        candidates.append(Path(sys.prefix) / "Library" / "bin" / "ffmpeg.exe")
    else:
        candidates.append(Path(sys.prefix) / "bin" / "ffmpeg")

    checked: set[Path] = set()
    for candidate in candidates:
        candidate = candidate.resolve()
        if candidate in checked or not candidate.is_file():
            continue
        checked.add(candidate)
        result = subprocess.run(
            [str(candidate), "-version"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        if result.returncode == 0:
            return candidate
    raise RuntimeError(
        "找不到可用的 ffmpeg。请在 fgo 环境安装 ffmpeg，或使用 --ffmpeg 指定路径。"
    )


def select_mp3_encoder(ffmpeg: Path) -> str | None:
    result = subprocess.run(
        [str(ffmpeg), "-hide_banner", "-encoders"],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError("无法读取 ffmpeg 编码器列表。")
    # 不使用 Windows 的 mp3_mf：它会截短数百毫秒的极短语音。
    for encoder in ("libmp3lame", "libshine"):
        if re.search(rf"\b{re.escape(encoder)}\b", result.stdout):
            return encoder
    return None


def find_lame(explicit: str | None) -> Path:
    candidates: list[Path] = []
    if explicit:
        candidates.append(Path(explicit).expanduser())
    found = shutil.which("lame")
    if found:
        candidates.append(Path(found))
    executable = "lame.exe" if os.name == "nt" else "lame"
    candidates.extend(
        [
            Path(sys.prefix) / "Library" / "bin" / executable,
            Path(sys.prefix) / "bin" / executable,
            Path(sys.prefix) / executable,
        ]
    )

    checked: set[Path] = set()
    for candidate in candidates:
        candidate = candidate.resolve()
        if candidate in checked or not candidate.is_file():
            continue
        checked.add(candidate)
        result = subprocess.run(
            [str(candidate), "--version"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        if result.returncode == 0:
            return candidate
    raise RuntimeError(
        "ffmpeg 没有 libmp3lame，且找不到 LAME。"
        "请在 fgo 环境安装 main::lame，或使用 --lame 指定路径。"
    )


def find_ffprobe(ffmpeg: Path) -> Path | None:
    executable = "ffprobe.exe" if os.name == "nt" else "ffprobe"
    sibling = ffmpeg.with_name(executable)
    if sibling.is_file():
        return sibling
    found = shutil.which("ffprobe")
    return Path(found).resolve() if found else None


def validate_mp3(path: Path, ffprobe: Path | None) -> None:
    if not path.is_file() or path.stat().st_size < 256:
        raise OSError(f"生成的 MP3 为空或过小：{path}")
    if ffprobe is None:
        return
    result = subprocess.run(
        [
            str(ffprobe),
            "-v",
            "error",
            "-select_streams",
            "a:0",
            "-show_entries",
            "stream=codec_name",
            "-of",
            "default=noprint_wrappers=1:nokey=1",
            str(path),
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    if result.returncode != 0 or "mp3" not in result.stdout.lower():
        detail = result.stderr.strip() or result.stdout.strip() or "未知错误"
        raise OSError(f"ffprobe 未确认文件为 MP3：{detail}")


def source_cache_path(cache_dir: Path, url: str) -> Path:
    suffix = Path(unquote(urlsplit(url).path)).suffix.lower()
    if not re.fullmatch(r"\.[a-z0-9]{1,6}", suffix):
        suffix = ".audio"
    digest = hashlib.sha256(url.encode("utf-8")).hexdigest()
    return cache_dir / "audio" / f"{digest}{suffix}"


def materialise_mp3(
    source: Path,
    source_url: str,
    target: Path,
    ffmpeg: Path,
    encoder: str | None,
    lame: Path | None,
    ffprobe: Path | None,
) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.stem}.{os.getpid()}.tmp.mp3")
    temporary_wav = target.with_name(f".{target.stem}.{os.getpid()}.tmp.wav")
    temporary.unlink(missing_ok=True)
    temporary_wav.unlink(missing_ok=True)
    source_suffix = Path(unquote(urlsplit(source_url).path)).suffix.lower()
    try:
        if source_suffix == ".mp3":
            shutil.copyfile(source, temporary)
        elif encoder is not None:
            command = [
                str(ffmpeg),
                "-nostdin",
                "-hide_banner",
                "-loglevel",
                "error",
                "-y",
                "-i",
                str(source),
                "-map_metadata",
                "-1",
                "-vn",
                "-c:a",
                encoder,
                "-b:a",
                "192k",
                str(temporary),
            ]
            result = subprocess.run(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                check=False,
            )
            if result.returncode != 0:
                detail = result.stderr.strip() or "未知 ffmpeg 错误"
                raise RuntimeError(f"ffmpeg 转码失败：{detail}")
        else:
            if lame is None:
                raise RuntimeError("没有可用的 LAME MP3 编码器。")
            # 先统一解码为 PCM WAV，使 LAME 同时支持 WAV、OGG 等源格式。
            decode = subprocess.run(
                [
                    str(ffmpeg),
                    "-nostdin",
                    "-hide_banner",
                    "-loglevel",
                    "error",
                    "-y",
                    "-i",
                    str(source),
                    "-map_metadata",
                    "-1",
                    "-vn",
                    "-c:a",
                    "pcm_s16le",
                    str(temporary_wav),
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                check=False,
            )
            if decode.returncode != 0:
                detail = decode.stderr.strip() or "未知 ffmpeg 错误"
                raise RuntimeError(f"ffmpeg 解码失败：{detail}")
            encode = subprocess.run(
                [str(lame), "--silent", "-b", "192", str(temporary_wav), str(temporary)],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                check=False,
            )
            if encode.returncode != 0:
                detail = encode.stderr.strip() or encode.stdout.strip() or "未知 LAME 错误"
                raise RuntimeError(f"LAME 编码失败：{detail}")
        validate_mp3(temporary, ffprobe)
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)
        temporary_wav.unlink(missing_ok=True)


def atomic_write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".part")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    os.replace(temporary, path)


def existing_mp3(path: Path) -> bool:
    return path.is_file() and path.stat().st_size >= 256


def process_servant(
    servant: Servant,
    index: int,
    total: int,
    args: argparse.Namespace,
    client: HttpClient,
    cache_dir: Path,
    ffmpeg: Path | None,
    encoder: str | None,
    lame: Path | None,
    ffprobe: Path | None,
) -> dict[str, object]:
    logging.info(
        "[%d/%d] %s / %s（Excel %s2）",
        index,
        total,
        servant.class_name,
        servant.name,
        servant.column,
    )
    page_url = voice_page_url(servant.name)
    page_html, cache_path, from_cache = load_page(
        client, cache_dir, servant.name, page_url, args.refresh_pages
    )
    clips = extract_voice_clips(page_html, page_url)

    # 旧缓存若是临时错误页，自动刷新一次。
    if not clips and from_cache:
        logging.warning("页面缓存中没有宝具语音，自动刷新：%s", servant.name)
        cache_path.unlink(missing_ok=True)
        page_html, cache_path, _ = load_page(
            client, cache_dir, servant.name, page_url, refresh=True
        )
        clips = extract_voice_clips(page_html, page_url)
    if not clips:
        raise RuntimeError(f"语音页没有找到“宝具卡/宝具”媒体行：{page_url}")

    logging.info(
        "找到 %d 条：%s", len(clips), "、".join(clip.filename for clip in clips)
    )
    servant_dir = (
        args.output_dir
        / safe_component(servant.class_name)
        / safe_component(servant.name)
    )
    downloaded: list[str] = []
    skipped: list[str] = []

    for clip in clips:
        target = servant_dir / clip.filename
        if existing_mp3(target) and not args.force:
            logging.info("跳过已有文件：%s", target)
            skipped.append(str(target))
            continue
        if args.dry_run:
            logging.info("[dry-run] %s <- %s", target, clip.url)
            continue
        if ffmpeg is None or (encoder is None and lame is None):
            raise AssertionError("下载模式必须准备 ffmpeg 和 MP3 编码器")

        raw_path = source_cache_path(cache_dir, clip.url)
        if not raw_path.is_file() or raw_path.stat().st_size == 0:
            logging.info("下载：%s", clip.url)
            client.download(clip.url, raw_path, page_url)
        else:
            logging.debug("复用源文件缓存：%s", raw_path)

        converted = False
        try:
            materialise_mp3(
                raw_path, clip.url, target, ffmpeg, encoder, lame, ffprobe
            )
            converted = True
            downloaded.append(str(target))
            logging.info("保存：%s", target)
        finally:
            if converted and not args.keep_source:
                raw_path.unlink(missing_ok=True)

    return {
        "status": "dry-run" if args.dry_run else "ok",
        "column": servant.column,
        "class": servant.class_name,
        "name": servant.name,
        "page_url": page_url,
        "page_cache": str(cache_path),
        "clips": [asdict(clip) for clip in clips],
        "downloaded": downloaded,
        "skipped": skipped,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="从 fgo.wiki 下载 Excel 中从者的宝具卡与宝具语音，并统一保存为 MP3。",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--sheet", help="工作表名；默认使用活动工作表")
    parser.add_argument("--start-column", default="E", help="从者区域起始列")
    parser.add_argument(
        "--servant",
        action="append",
        help="只处理指定正式从者名；可以多次提供",
    )
    parser.add_argument("--start-from", help="从指定正式从者名开始处理")
    parser.add_argument("--limit", type=int, help="最多处理多少名从者")
    parser.add_argument("--list", action="store_true", help="只列出 Excel 解析结果")
    parser.add_argument("--dry-run", action="store_true", help="解析页面但不下载音频")
    parser.add_argument("--force", action="store_true", help="覆盖已有 MP3")
    parser.add_argument(
        "--refresh-pages", action="store_true", help="忽略 HTML 页面缓存"
    )
    parser.add_argument(
        "--keep-source", action="store_true", help="转码成功后保留原始音频缓存"
    )
    parser.add_argument("--ffmpeg", help="ffmpeg 可执行文件路径")
    parser.add_argument("--lame", help="LAME 可执行文件路径")
    parser.add_argument("--timeout", type=float, default=90.0, help="单次请求超时秒数")
    parser.add_argument("--retries", type=int, default=5, help="失败后的最大重试次数")
    parser.add_argument("--delay", type=float, default=0.5, help="两次网络请求最小间隔秒数")
    parser.add_argument("--verbose", action="store_true")
    return parser


def select_servants(servants: list[Servant], args: argparse.Namespace) -> list[Servant]:
    selected = servants
    if args.servant:
        wanted = set(args.servant)
        available = {servant.name for servant in servants}
        missing = sorted(wanted - available)
        if missing:
            raise ValueError(f"Excel 中找不到指定从者：{'、'.join(missing)}")
        selected = [servant for servant in selected if servant.name in wanted]

    if args.start_from:
        try:
            start = next(
                index
                for index, servant in enumerate(selected)
                if servant.name == args.start_from
            )
        except StopIteration as exc:
            raise ValueError(f"选中列表中找不到 --start-from：{args.start_from}") from exc
        selected = selected[start:]

    if args.limit is not None:
        if args.limit < 1:
            raise ValueError("--limit 必须大于 0")
        selected = selected[: args.limit]
    return selected


def main() -> int:
    configure_console()
    parser = build_parser()
    args = parser.parse_args()
    configure_logging(args.output_dir, args.verbose)

    try:
        servants = read_servants(args.workbook, args.sheet, args.start_column)
        logging.info("从 Excel 读取到 %d 名从者。", len(servants))
        selected = select_servants(servants, args)
        logging.info("本次选择 %d 名从者。", len(selected))

        if args.list:
            for servant in selected:
                print(
                    f"{servant.column}2\t{servant.class_name}\t{servant.name}"
                )
            return 0

        ffmpeg: Path | None = None
        encoder: str | None = None
        lame: Path | None = None
        ffprobe: Path | None = None
        if not args.dry_run:
            ffmpeg = find_ffmpeg(args.ffmpeg)
            encoder = select_mp3_encoder(ffmpeg)
            if encoder is None:
                lame = find_lame(args.lame)
            ffprobe = find_ffprobe(ffmpeg)
            logging.info(
                "ffmpeg：%s；MP3 编码器：%s",
                ffmpeg,
                encoder or f"LAME CLI ({lame})",
            )

        client = HttpClient(args.timeout, args.retries, args.delay)
        cache_dir = args.output_dir / ".cache"
        report_path = args.output_dir / "download_report.json"
        report: dict[str, object] = {
            "workbook": str(args.workbook.resolve()),
            "output_dir": str(args.output_dir.resolve()),
            "total_in_workbook": len(servants),
            "selected": len(selected),
            "started_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "results": [],
        }
        failures = 0

        for index, servant in enumerate(selected, start=1):
            try:
                result = process_servant(
                    servant,
                    index,
                    len(selected),
                    args,
                    client,
                    cache_dir,
                    ffmpeg,
                    encoder,
                    lame,
                    ffprobe,
                )
            except KeyboardInterrupt:
                raise
            except Exception as exc:
                failures += 1
                logging.exception("处理失败：%s / %s", servant.class_name, servant.name)
                result = {
                    "status": "error",
                    "column": servant.column,
                    "class": servant.class_name,
                    "name": servant.name,
                    "page_url": voice_page_url(servant.name),
                    "error": f"{type(exc).__name__}: {exc}",
                }
            results = report["results"]
            assert isinstance(results, list)
            results.append(result)
            atomic_write_json(report_path, report)

        report["finished_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
        report["failures"] = failures
        atomic_write_json(report_path, report)
        if failures:
            logging.error(
                "完成，但有 %d 名从者失败。详情见 %s", failures, report_path
            )
            return 1
        logging.info("全部完成。报告：%s", report_path)
        return 0
    except KeyboardInterrupt:
        logging.error("用户中断。已完成文件不会丢失，下次运行会自动续传。")
        return 130
    except Exception:
        logging.exception("程序无法继续")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
